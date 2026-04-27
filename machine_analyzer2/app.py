"""
STARTECH — app.py
INF232 · Statistique Descriptive · Université de Yaoundé 1
Backend Flask : auth, CRUD, régression multiple, upload image, export Excel
"""

import io
import os
import re
import numpy as np
import pandas as pd
from functools import wraps
from flask import (
    Flask, render_template, request, jsonify,
    make_response, session, redirect, url_for, send_file
)
from flask_cors import CORS
from werkzeug.utils import secure_filename
from database import (
    init_db, init_default_users,
    add_machine, get_all_machines, get_stats,
    delete_machine, update_machine, get_machine_by_id,
    update_machine_image, get_user_by_username,
    bulk_insert
)

# ── Configuration ─────────────────────────────────────────────────────────────
BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER   = os.path.join(BASE_DIR, 'static', 'uploads')
ALLOWED_EXT     = {'jpg', 'jpeg', 'png', 'webp'}
MAX_UPLOAD_SIZE = 5 * 1024 * 1024   # 5 MB

app = Flask(__name__)
app.secret_key            = os.environ.get('SECRET_KEY', 'startech-inf232-yaounde-2024')
app.config['UPLOAD_FOLDER']      = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_SIZE
CORS(app, supports_credentials=True)

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
init_db()
init_default_users()


# ── Helpers ───────────────────────────────────────────────────────────────────
def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT


def extract_gen_number(gen_str: str) -> int:
    """Extrait le numéro de génération d'une chaîne comme '12ème gén'."""
    if not gen_str:
        return 0
    m = re.search(r'\d+', str(gen_str))
    return int(m.group()) if m else 0


def _regression_data(machines):
    """Construit les arrays numpy pour la régression multiple."""
    valid = []
    for m in machines:
        if not (m.get('price') and m['price'] > 0 and m.get('ram') and m['ram'] > 0):
            continue
        valid.append((
            float(m.get('ram',        0) or 0),
            float(m.get('storage',    0) or 0),
            float(m.get('gpu_memory', 0) or 0),
            float(extract_gen_number(m.get('cpu_gen', ''))),
            float(m['price'])
        ))
    return valid


# ── Décorateurs d'authentification ────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Non authentifié'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            return jsonify({'error': 'Accès refusé — rôle admin requis'}), 403
        return f(*args, **kwargs)
    return decorated


# ════════════════════════════════════════════════════════════════════════════════
#  PAGES
# ════════════════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    if 'user' in session:
        return redirect(url_for('admin_page') if session['role'] == 'admin' else url_for('client_page'))
    return redirect(url_for('login_page'))


@app.route('/login')
def login_page():
    if 'user' in session:
        return redirect(url_for('admin_page') if session['role'] == 'admin' else url_for('client_page'))
    return render_template('login.html')


@app.route('/client')
@login_required
def client_page():
    return render_template('client.html',
                           username=session.get('user'),
                           role=session.get('role'))


@app.route('/admin')
@login_required
def admin_page():
    if session.get('role') != 'admin':
        return redirect(url_for('client_page'))
    return render_template('admin.html', username=session.get('user'))


# ════════════════════════════════════════════════════════════════════════════════
#  AUTH API
# ════════════════════════════════════════════════════════════════════════════════

@app.route('/api/auth/login', methods=['POST'])
def api_login():
    data     = request.json or {}
    username = data.get('username', '').strip()
    password = data.get('password', '')

    if not username or not password:
        return jsonify({'error': 'Identifiants manquants'}), 400

    user = get_user_by_username(username)
    if not user:
        return jsonify({'error': 'Identifiants incorrects'}), 401

    from werkzeug.security import check_password_hash
    if not check_password_hash(user['password_hash'], password):
        return jsonify({'error': 'Identifiants incorrects'}), 401

    session.permanent  = True
    session['user']    = user['username']
    session['role']    = user['role']

    redirect_url = '/admin' if user['role'] == 'admin' else '/client'
    return jsonify({'message': 'Connexion réussie', 'role': user['role'], 'redirect': redirect_url})


@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'message': 'Déconnecté'})


@app.route('/api/auth/me')
def api_me():
    if 'user' not in session:
        return jsonify({'authenticated': False}), 401
    return jsonify({'authenticated': True, 'user': session['user'], 'role': session['role']})


# ════════════════════════════════════════════════════════════════════════════════
#  MACHINES — CRUD
# ════════════════════════════════════════════════════════════════════════════════

@app.route('/api/machines', methods=['GET'])
@login_required
def list_machines():
    return jsonify(get_all_machines())


@app.route('/api/machines', methods=['POST'])
@login_required
@admin_required
def create_machine():
    data = request.json or {}
    try:
        add_machine(
            data['name'], data['brand'],
            data.get('cpu', ''),      data.get('cpu_gen', ''),
            float(data.get('cpu_speed',  0) or 0),
            int(  data.get('ram',        0) or 0),
            int(  data.get('storage',    0) or 0),
            int(  data.get('gpu_memory', 0) or 0),
            float(data['price']),
            data.get('status',        'En stock'),
            data.get('serial_number', ''),
            data.get('location',      ''),
            data.get('notes',         ''),
            int(  data.get('quantity',   1) or 1)
        )
        return jsonify({'message': 'Machine ajoutée avec succès'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/machines/<int:mid>', methods=['GET'])
@login_required
def get_machine(mid):
    m = get_machine_by_id(mid)
    return jsonify(m) if m else (jsonify({'error': 'Non trouvée'}), 404)


@app.route('/api/machines/<int:mid>', methods=['PUT'])
@login_required
@admin_required
def modify_machine(mid):
    if not get_machine_by_id(mid):
        return jsonify({'error': 'Non trouvée'}), 404
    data = request.json or {}
    try:
        update_machine(
            mid,
            data['name'], data['brand'],
            data.get('cpu', ''),      data.get('cpu_gen', ''),
            float(data.get('cpu_speed',  0) or 0),
            int(  data.get('ram',        0) or 0),
            int(  data.get('storage',    0) or 0),
            int(  data.get('gpu_memory', 0) or 0),
            float(data['price']),
            data.get('status',        'En stock'),
            data.get('serial_number', ''),
            data.get('location',      ''),
            data.get('notes',         ''),
            int(  data.get('quantity',   1) or 1)
        )
        return jsonify({'message': 'Machine mise à jour'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/machines/<int:mid>', methods=['DELETE'])
@login_required
@admin_required
def remove_machine(mid):
    m = get_machine_by_id(mid)
    if not m:
        return jsonify({'error': 'Non trouvée'}), 404
    # Suppression de l'image associée
    if m.get('image_path'):
        try:
            img_path = os.path.join(BASE_DIR, 'static', 'uploads',
                                    os.path.basename(m['image_path']))
            if os.path.exists(img_path):
                os.remove(img_path)
        except Exception:
            pass
    delete_machine(mid)
    return jsonify({'message': 'Machine supprimée'})


# ── Upload image ──────────────────────────────────────────────────────────────

@app.route('/api/machines/<int:mid>/image', methods=['POST'])
@login_required
@admin_required
def upload_image(mid):
    m = get_machine_by_id(mid)
    if not m:
        return jsonify({'error': 'Machine non trouvée'}), 404

    if 'image' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400

    file = request.files['image']
    if not file or file.filename == '':
        return jsonify({'error': 'Fichier vide'}), 400

    if not allowed_file(file.filename):
        return jsonify({'error': 'Format non autorisé (jpg, png, webp)'}), 400

    # Supprimer l'ancienne image
    if m.get('image_path'):
        old = os.path.join(BASE_DIR, 'static', 'uploads',
                           os.path.basename(m['image_path']))
        if os.path.exists(old):
            try:
                os.remove(old)
            except Exception:
                pass

    ext      = file.filename.rsplit('.', 1)[1].lower()
    filename = secure_filename(f'machine_{mid}.{ext}')
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    image_url = f'/static/uploads/{filename}'
    update_machine_image(mid, image_url)
    return jsonify({'message': 'Image téléchargée', 'image_path': image_url})


@app.route('/api/machines/<int:mid>/image', methods=['DELETE'])
@login_required
@admin_required
def delete_image(mid):
    m = get_machine_by_id(mid)
    if not m:
        return jsonify({'error': 'Non trouvée'}), 404
    if m.get('image_path'):
        old = os.path.join(BASE_DIR, 'static', 'uploads',
                           os.path.basename(m['image_path']))
        if os.path.exists(old):
            os.remove(old)
        update_machine_image(mid, None)
    return jsonify({'message': 'Image supprimée'})


# ════════════════════════════════════════════════════════════════════════════════
#  STATS
# ════════════════════════════════════════════════════════════════════════════════

@app.route('/api/stats')
@login_required
def stats():
    return jsonify(get_stats())


# ════════════════════════════════════════════════════════════════════════════════
#  RÉGRESSION
# ════════════════════════════════════════════════════════════════════════════════

@app.route('/api/regression/multiple')
@login_required
def multiple_regression():
    """Régression linéaire multiple : Prix = a1·RAM + a2·Stockage + a3·GPU + a4·Génération + b"""
    valid = _regression_data(get_all_machines())

    if len(valid) < 5:
        return jsonify({'error': 'Pas assez de données (minimum 5 machines)'}), 400

    arr = np.array(valid, dtype=float)
    X   = np.column_stack([arr[:, :4], np.ones(len(arr))])
    y   = arr[:, 4]

    coeffs, _, _, _ = np.linalg.lstsq(X, y, rcond=None)
    y_pred = X @ coeffs
    ss_res = float(np.sum((y - y_pred) ** 2))
    ss_tot = float(np.sum((y - np.mean(y)) ** 2))
    r2     = 1.0 - ss_res / ss_tot if ss_tot else 0.0

    a_ram, a_sto, a_gpu, a_gen, b = [float(c) for c in coeffs]

    scatter = [{
        'ram': v[0], 'storage': v[1], 'gpu': v[2], 'gen': v[3],
        'price': v[4],
        'predicted': float(a_ram*v[0] + a_sto*v[1] + a_gpu*v[2] + a_gen*v[3] + b)
    } for v in valid]

    return jsonify({
        'a_ram':     round(a_ram, 2),
        'a_storage': round(a_sto, 2),
        'a_gpu':     round(a_gpu, 2),
        'a_gen':     round(a_gen, 2),
        'b':         round(b, 2),
        'r2':        round(r2, 4),
        'n':         len(valid),
        'scatter':   scatter,
        'formula':   (f'Prix = {round(a_ram):,}·RAM + {round(a_sto):,}·Stockage '
                      f'+ {round(a_gpu):,}·GPU + {round(a_gen):,}·Génération + {round(b):,}')
    })


@app.route('/api/regression/simple')
@login_required
def simple_regressions():
    """4 régressions simples indépendantes : chaque variable vs Prix"""
    machines = get_all_machines()

    extractors = {
        'ram':        ('RAM (Go)',        lambda m: float(m.get('ram', 0) or 0)),
        'storage':    ('Stockage (Go)',   lambda m: float(m.get('storage', 0) or 0)),
        'gpu_memory': ('GPU dédié (Go)',  lambda m: float(m.get('gpu_memory', 0) or 0)),
        'cpu_gen':    ('Génération CPU',  lambda m: float(extract_gen_number(m.get('cpu_gen', '')))),
    }

    results = {}
    for key, (label, fn) in extractors.items():
        pairs = [(fn(m), float(m['price'])) for m in machines
                 if m.get('price') and m['price'] > 0 and fn(m) > 0]

        if len(pairs) < 3:
            results[key] = {'error': 'Pas assez de données', 'label': label}
            continue

        xs   = np.array([p[0] for p in pairs])
        ys   = np.array([p[1] for p in pairs])
        a, b = np.polyfit(xs, ys, 1)

        yp    = a * xs + b
        ss_r  = float(np.sum((ys - yp) ** 2))
        ss_t  = float(np.sum((ys - np.mean(ys)) ** 2))
        r2    = 1.0 - ss_r / ss_t if ss_t else 0.0

        x_line = sorted(set(xs.tolist()))
        results[key] = {
            'label':   label,
            'a':       round(float(a), 2),
            'b':       round(float(b), 2),
            'r2':      round(r2, 4),
            'n':       len(pairs),
            'scatter': [{'x': float(x), 'y': float(y)} for x, y in pairs],
            'line':    [{'x': x, 'y': float(a * x + b)} for x in x_line],
        }

    return jsonify(results)


@app.route('/api/predict/multiple')
@login_required
def predict_multiple():
    """Prédiction par régression multiple."""
    try:
        ram     = float(request.args.get('ram',     0) or 0)
        storage = float(request.args.get('storage', 0) or 0)
        gpu     = float(request.args.get('gpu',     0) or 0)
        gen     = float(request.args.get('gen',     0) or 0)
    except ValueError:
        return jsonify({'error': 'Paramètres invalides'}), 400

    valid = _regression_data(get_all_machines())
    if len(valid) < 5:
        return jsonify({'error': 'Pas assez de données'}), 400

    arr    = np.array(valid, dtype=float)
    X      = np.column_stack([arr[:, :4], np.ones(len(arr))])
    y      = arr[:, 4]
    coeffs, _, _, _ = np.linalg.lstsq(X, y, rcond=None)

    predicted = max(0.0, float(np.dot(coeffs, [ram, storage, gpu, gen, 1])))

    yp    = X @ coeffs
    ss_r  = float(np.sum((y - yp) ** 2))
    ss_t  = float(np.sum((y - np.mean(y)) ** 2))
    r2    = 1.0 - ss_r / ss_t if ss_t else 0.0

    return jsonify({
        'predicted_price': round(predicted),
        'r2':     round(r2, 4),
        'formula': (f'Prix = {round(float(coeffs[0])):,}·RAM + {round(float(coeffs[1])):,}·Stockage '
                    f'+ {round(float(coeffs[2])):,}·GPU + {round(float(coeffs[3])):,}·Génération '
                    f'+ {round(float(coeffs[4])):,}')
    })


# ════════════════════════════════════════════════════════════════════════════════
#  EXPORT EXCEL (pandas + openpyxl)
# ════════════════════════════════════════════════════════════════════════════════

@app.route('/api/export/excel')
@login_required
@admin_required
def export_excel():
    machines = get_all_machines()
    df       = pd.DataFrame(machines)

    rename_map = {
        'id': 'ID', 'name': 'Modèle', 'brand': 'Marque',
        'cpu': 'CPU', 'cpu_gen': 'Génération CPU', 'cpu_speed': 'Fréq. CPU (GHz)',
        'ram': 'RAM (Go)', 'storage': 'Stockage (Go)', 'gpu_memory': 'GPU (Go)',
        'price': 'Prix (FCFA)', 'status': 'Statut',
        'serial_number': 'N° Série', 'location': 'Emplacement',
        'notes': 'Notes', 'created_at': 'Date ajout', 'image_path': 'Photo'
    }
    df = df.rename(columns={k: v for k, v in rename_map.items() if k in df.columns})

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Inventaire STARTECH', index=False)
        ws = writer.sheets['Inventaire STARTECH']
        # Style header
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill('solid', fgColor='1E3A8A')
        for cell in ws[1]:
            cell.font      = Font(bold=True, color='FFFFFF', size=11)
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center')
        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    output.seek(0)
    return send_file(
        output,
        download_name='inventaire_startech.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ════════════════════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    app.run(debug=True, port=5000)
